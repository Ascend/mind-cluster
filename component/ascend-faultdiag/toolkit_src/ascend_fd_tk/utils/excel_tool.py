#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright 2026 Huawei Technologies Co., Ltd
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ==============================================================================

import os
from enum import Enum
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from . import logger

_CONSOLE_LOGGER = logger.CONSOLE_LOGGER


class Color(Enum):
    """颜色枚举类"""

    # 基本颜色
    RED = "FF0000"
    GREEN = "00FF00"
    BLUE = "0000FF"
    YELLOW = "FFFF00"
    CYAN = "00FFFF"
    MAGENTA = "FF00FF"
    WHITE = "FFFFFF"
    BLACK = "000000"

    # 常用颜色
    LIGHT_GRAY = "C0C0C0"
    GRAY = "808080"
    DARK_GRAY = "404040"
    LIGHT_BLUE = "B4C6E7"  # 蓝色 个性 1 淡色 60%
    LIGHT_ORANGE = "F8CBAD"  # 橙色 个性 6 淡色 60%
    LIGHT_TEAL = "B7DEE8"  # 水绿色 个性 5 淡色 60%

    # 业务相关颜色
    SUCCESS = "00FF00"  # 成功 - 绿色
    WARNING = "FFFF00"  # 警告 - 黄色
    ERROR = "FF0000"  # 错误 - 红色
    INFO = "0000FF"  # 信息 - 蓝色

    # 柔和的业务相关颜色
    LIGHT_SUCCESS = "E6F9E6"  # 柔和成功 - 浅绿色
    LIGHT_WARNING = "FFF3CD"  # 柔和警告 - 浅黄色
    LIGHT_ERROR = "F8D7DA"  # 柔和错误 - 浅红色


_COLOR_LIST = [Color.LIGHT_BLUE.value, Color.LIGHT_ORANGE.value, Color.LIGHT_TEAL.value]


class CellStyle:
    """单元格样式类"""

    def __init__(self, bg_color: Optional[Color] = None, font_color: Optional[Color] = None):
        """
        初始化单元格样式
        :param bg_color: 背景颜色
        :param font_color: 字体颜色
        """
        self.bg_color = bg_color
        self.font_color = font_color


class StyledCell:
    """带样式的单元格类"""

    def __init__(self, value: Any, style: Optional[CellStyle] = None):
        """
        初始化带样式的单元格
        :param value: 单元格值
        :param style: 单元格样式
        """
        self.value = value
        self.style = style or CellStyle()


def flatten_dict(d: Dict[str, Any], parent_key: str = '', sep: str = '_') -> Dict[str, Any]:
    """
    递归展平嵌套字典（如 {'a': {'b': 1}} -> {'a_b': 1}）
    :param d: 可能包含嵌套结构的字典
    :param parent_key: 父键名（用于拼接嵌套键）
    :param sep: 键名分隔符
    :return: 展平后的字典
    """
    res: List[tuple] = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict) and not isinstance(v, StyledCell):  # 不展平StyledCell对象
            # 递归展平嵌套字典
            res.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            res.append((new_key, v))
    return dict(res)


def get_all_columns(dict_list: List[Dict[str, Any]], flatten: bool = True, sep: str = '_') -> List[str]:
    """
    获取所有字典中的键（列名），自动去重并排序
    :param dict_list: 字典列表
    :param flatten: 是否展平嵌套字典
    :param sep: 嵌套键分隔符
    :return: 所有列名的列表
    """
    columns = set()
    for d in dict_list:
        if flatten:
            # 展平后获取键
            flat_d = flatten_dict(d, sep=sep)
            columns.update(flat_d.keys())
        else:
            # 不展平，直接获取顶层键
            columns.update(d.keys())
    # 排序列名，确保输出顺序一致
    return sorted(columns)


class ExcelGenerator:
    """
    Excel文件生成器，支持将多个List[dict]数据写入不同的sheet
    """

    DEFAULT_BORDER = "thin"
    DEFAULT_COLOR = "000000"

    def __init__(self):
        """初始化Excel生成器"""
        self.workbook = Workbook()
        # 删除默认创建的sheet
        if self.workbook.active:
            self.workbook.remove(self.workbook.active)
        self.sheets_data = []

        # 定义默认边框样式
        self.default_border = Border(
            left=Side(border_style=self.DEFAULT_BORDER, color=self.DEFAULT_COLOR),
            right=Side(border_style=self.DEFAULT_BORDER, color=self.DEFAULT_COLOR),
            top=Side(border_style=self.DEFAULT_BORDER, color=self.DEFAULT_COLOR),
            bottom=Side(border_style=self.DEFAULT_BORDER, color=self.DEFAULT_COLOR),
        )

    @staticmethod
    def _merge_cells(sheet, sheet_info):
        # 处理精确指定的合并范围
        for merge_range in sheet_info.get('merge_cells', []):
            min_row, min_col, max_row, max_col = merge_range
            sheet.merge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col,
            )

    @staticmethod
    def _get_cell_value_and_style(cell_data, na_rep: str):
        """
        提取单元格值和样式
        :param cell_data: 单元格数据，可以是简单值或StyledCell对象
        :param na_rep: 空值替换字符串
        :return: (值, CellStyle对象)
        """
        if cell_data is None:
            return na_rep, CellStyle()

        if isinstance(cell_data, StyledCell):
            return cell_data.value, cell_data.style

        return cell_data, CellStyle()

    def add_sheet(
        self,
        sheet_name: str,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None,
        flatten: bool = False,
        sep: str = '_',
        na_rep: str = '',
        header_widths: Optional[Dict[str, int]] = None,
        merged_headers: Optional[List[Dict[str, List[str]]]] = None,
        merge_cells: Optional[List[Tuple[int, int, int, int]]] = None,
        tab_color: Optional[str] = None,
    ):
        """
        添加一个sheet的数据
        :param sheet_name: sheet名称
        :param data: 数据列表，每个元素为一个字典
        :param columns: 自定义列顺序（None 则自动获取所有列）
        :param flatten: 是否展平嵌套字典（如 {'a': {'b': 1}} -> 'a_b' 列）
        :param sep: 嵌套键的连接符（展平时使用）
        :param na_rep: 空值替换字符串（默认空字符串）
        :param header_widths: 自定义列宽，格式为 {列名: 宽度}，优先级高于自动调整
        :param merged_headers: 合并的header配置列表，每个元素为 {合并单元格内容: [要合并的列名列表]}
        :param merge_cells: 单元格合并范围列表，每个元素为 (min_row, min_col, max_row, max_col)，
                            行号从1开始(1=表头)，列号从1开始
        :param tab_color: Sheet标签颜色，十六进制颜色值，如 "4472C4"
        """
        if not data:
            raise ValueError("输入的数据列表不能为空")

        # 处理嵌套字典（展平）
        processed_data = []
        for d in data:
            if flatten:
                processed_data.append(flatten_dict(d, sep=sep))
            else:
                processed_data.append(d.copy())

        # 确定列名（自定义列或自动获取）
        if columns is None:
            columns = get_all_columns(processed_data, flatten=False)  # 已展平，无需再次处理
        else:
            # 检查自定义列是否存在
            all_cols = get_all_columns(processed_data, flatten=False)
            invalid_cols = [col for col in columns if col not in all_cols]
            if invalid_cols:
                _CONSOLE_LOGGER.info("警告：自定义列中存在不存在的键：%s（将输出空值）", str(invalid_cols))

        # 保存sheet数据
        self.sheets_data.append(
            {
                'sheet_name': sheet_name,
                'data': processed_data,
                'columns': columns,
                'na_rep': na_rep,
                'header_widths': header_widths or {},
                'merged_headers': merged_headers or [],
                'merge_cells': merge_cells or [],
                'tab_color': tab_color,
            }
        )

    def generate_excel(self, output_path: str):
        """
        生成Excel文件
        :param output_path: 输出Excel文件路径
        """
        if not self.sheets_data:
            raise ValueError("没有添加任何sheet数据")

        # 确保输出目录存在
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        # 为每个sheet添加数据
        for sheet_info in self.sheets_data:
            sheet_name = sheet_info['sheet_name']
            data = sheet_info['data']
            columns = sheet_info['columns']
            na_rep = sheet_info['na_rep']
            header_widths = sheet_info['header_widths']

            # 创建sheet
            sheet = self.workbook.create_sheet(title=sheet_name)

            # 设置Sheet标签颜色
            tab_color = sheet_info.get('tab_color')
            if tab_color:
                sheet.sheet_properties.tabColor = tab_color

            # 写入合并的header（第一行开始）
            merged_headers = sheet_info.get('merged_headers', [])
            header_row = 1
            if merged_headers:
                header_row = len(merged_headers) + 1
                self._merge_header(sheet, merged_headers, columns)

            # 写入表头
            for col_idx, col_name in enumerate(columns, 1):
                cell = sheet.cell(row=header_row, column=col_idx, value=col_name)
                # 设置表头样式
                cell.font = Font(bold=True)
                # 添加边框
                cell.border = self.default_border

            # 写入数据行
            data_start_row = header_row + 1
            sheet_info['data_start_row'] = data_start_row
            for row_idx, row_data in enumerate(data, data_start_row):
                for col_idx, col_name in enumerate(columns, 1):
                    cell_data = row_data.get(col_name, na_rep)
                    value, style = self._get_cell_value_and_style(cell_data, na_rep)

                    # 设置单元格值
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    # 添加边框
                    cell.border = self.default_border
                    # 设置对齐方式
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                    # 应用单元格样式
                    if style:
                        # 设置背景颜色
                        if style.bg_color:
                            cell.fill = PatternFill(
                                start_color=style.bg_color.value, end_color=style.bg_color.value, fill_type="solid"
                            )
                        # 设置字体颜色
                        if style.font_color:
                            # 保留原有的字体设置（如粗体），仅更新颜色
                            current_font = cell.font or Font()
                            cell.font = Font(
                                name=current_font.name,
                                size=current_font.size,
                                bold=current_font.bold,
                                italic=current_font.italic,
                                underline=current_font.underline,
                                color=style.font_color.value,
                            )

            self._merge_cells(sheet, sheet_info)

            # 设置列宽
            for col_idx in range(1, len(columns) + 1):
                column_letter = get_column_letter(col_idx)
                column_name = columns[col_idx - 1]

                # 优先使用自定义列宽
                if column_name in header_widths:
                    sheet.column_dimensions[column_letter].width = header_widths[column_name]
                else:
                    # 自动调整列宽
                    max_length = 0
                    # 检查表头长度
                    max_length = max(max_length, len(str(column_name)))
                    for row_idx in range(data_start_row, data_start_row + len(data)):
                        cell_value = str(sheet[f"{column_letter}{row_idx}"].value)
                        max_length = max(max_length, len(cell_value))
                    # 设置列宽（加一点边距）
                    sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # 保存Excel文件
        try:
            self.workbook.save(output_path)
            _CONSOLE_LOGGER.info("Excel文件已生成：%s（共 %d 个sheet）", output_path, len(self.sheets_data))
        except OSError as e:
            raise OSError(f"生成Excel文件到：{output_path} 失败，可能是已打开文件占用，异常：{e}") from e
        finally:
            # 关闭工作簿
            self.workbook.close()

    def _merge_header(self, sheet, merged_headers, columns):
        for header_row_idx, header_config in enumerate(merged_headers, 1):
            for idx, (header_text, column_names) in enumerate(header_config.items()):
                # 找到这些列在columns中的索引
                col_indices = [columns.index(col) + 1 for col in column_names if col in columns]
                if col_indices:
                    # 合并单元格
                    start_col = min(col_indices)
                    end_col = max(col_indices)
                    sheet.merge_cells(
                        start_row=header_row_idx, start_column=start_col, end_row=header_row_idx, end_column=end_col
                    )
                    # 设置合并单元格的值和样式
                    cell = sheet.cell(row=header_row_idx, column=start_col, value=header_text)
                    cell.font = Font(bold=True)
                    cell.border = self.default_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    color_idx = idx % len(_COLOR_LIST)
                    cell.fill = PatternFill(
                        fill_type='solid', start_color=_COLOR_LIST[color_idx], end_color=_COLOR_LIST[color_idx]
                    )


def dict_list_to_excel(
    dict_list: List[Dict[str, Any]],
    output_path: str,
    sheet_name: str = "Sheet1",
    columns: Optional[List[str]] = None,
    flatten: bool = False,
    sep: str = '_',
    na_rep: str = '',
    header_widths: Optional[Dict[str, int]] = None,
) -> None:
    """
    便捷函数：将单个List[dict]转换为Excel文件
    :param dict_list: 输入的字典列表（每个字典代表一行）
    :param output_path: 输出Excel文件路径
    :param sheet_name: sheet名称
    :param columns: 自定义列顺序（None 则自动获取所有列）
    :param flatten: 是否展平嵌套字典（如 {'a': {'b': 1}} -> 'a_b' 列）
    :param sep: 嵌套键的连接符（展平时使用）
    :param na_rep: 空值替换字符串（默认空字符串）
    :param header_widths: 自定义列宽，格式为 {列名: 宽度}，优先级高于自动调整
    """
    excel_gen = ExcelGenerator()
    excel_gen.add_sheet(sheet_name, dict_list, columns, flatten, sep, na_rep, header_widths)
    excel_gen.generate_excel(output_path)
