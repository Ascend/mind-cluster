#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright 2026 Huawei Technologies Co., Ltd
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ==============================================================================

from dataclasses import dataclass, field
import os
from typing import Dict, Optional, Tuple

from openpyxl import load_workbook

from ascend_fd_tk.core.model.host import HostInfo
from ascend_fd_tk.core.model.switch import SwitchInfo
from ascend_fd_tk.utils.logger import DIAG_LOGGER


@dataclass
class HostLocationInfo:
    host_name: str = ""
    host_sn: str = ""
    room_name: str = ""
    cabinet_id: str = ""
    l1_switch_name: str = ""
    l1_switch_ip: str = ""
    l1_switch_sn: str = ""


@dataclass
class SwitchLocationInfo:
    sn: str = ""
    room_name: str = ""
    cabinet_id: str = ""
    switch_name: str = ""
    switch_ip: str = ""


@dataclass
class LocationConfig:
    host_name_location_map: Dict[str, HostLocationInfo] = field(default_factory=dict)
    host_sn_location_map: Dict[str, HostLocationInfo] = field(default_factory=dict)

    switch_sn_location_map: Dict[str, SwitchLocationInfo] = field(default_factory=dict)
    switch_ip_location_map: Dict[str, SwitchLocationInfo] = field(default_factory=dict)
    switch_name_location_map: Dict[str, SwitchLocationInfo] = field(default_factory=dict)

    def enrich_host_info(self, host_info: HostInfo):
        if host_info:
            host_info.room_name, host_info.cabinet_id = self._find_host_location(host_info.sn_num, host_info.hostname)

    def enrich_switch_info(self, switch_info: SwitchInfo):
        if switch_info:
            switch_info.room_name, switch_info.cabinet_id = self._find_switch_location(
                switch_info.sn, switch_info.swi_id, switch_info.name
            )

    def _find_host_location(self, input_sn: str, input_name: str) -> Tuple[str, str]:
        """查找主机位置信息，查找顺序: SN -> name"""
        info = self.host_sn_location_map.get(input_sn) or self.host_name_location_map.get(input_name)
        if info:
            return info.room_name, info.cabinet_id
        return "", ""

    def _find_switch_location(self, input_sn: str, input_ip: str, input_name: str) -> Tuple[str, str]:
        """查找交换机位置信息，查找顺序: SN -> IP -> name"""
        info = (
            self.switch_sn_location_map.get(input_sn)
            or self.switch_ip_location_map.get(input_ip)
            or self.switch_name_location_map.get(input_name)
        )
        if info:
            return info.room_name, info.cabinet_id
        return "", ""


class LldConfigReader:
    LLD_FILE_NAME = "LLD.xlsx"
    L1_SHEET_NAME = "灵衢L1网络对应关系"
    L2_SHEET_NAME = "灵衢L2网络对应关系"

    @staticmethod
    def _build_col_index(header: list, expected_cols: list) -> Dict[str, int]:
        col_idx = {}
        for col_name in expected_cols:
            for i, h in enumerate(header):
                if h == col_name:
                    col_idx[col_name] = i
                    break
        return col_idx

    @staticmethod
    def _find_sheet_by_keyword(wb, keyword: str):
        for name in wb.sheetnames:
            if keyword in name:
                return wb[name]
        return None

    def read_from_dir(self, dir_path: str) -> Optional[LocationConfig]:
        file_path = os.path.join(dir_path, self.LLD_FILE_NAME)
        if not os.path.exists(file_path):
            DIAG_LOGGER.warning("目录%s中未找到%s配置文件", dir_path, self.LLD_FILE_NAME)
            return None
        DIAG_LOGGER.info("读取配置文件: %s", file_path)
        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        except Exception as e:
            DIAG_LOGGER.error("读取配置文件%s失败: %s", file_path, e)
            return None
        try:
            return self._read(wb)
        finally:
            wb.close()

    def _read(self, wb) -> LocationConfig:
        config = LocationConfig()
        self._read_l1_sheet(wb, config)
        self._read_l2_sheet(wb, config)
        host_count = len(config.host_name_location_map)
        switch_count = len(config.switch_name_location_map)
        DIAG_LOGGER.info("加载 %d 条服务器信息，%d 条交换机信息", host_count, switch_count)
        return config

    def _iter_sheet_rows(self, wb, sheet_name: str, expected_cols: list):
        """迭代 Excel sheet 的数据行，yield (col_idx, values)"""
        ws = self._find_sheet_by_keyword(wb, sheet_name)
        if not ws:
            DIAG_LOGGER.warning("未找到包含%s的Sheet", sheet_name)
            return
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:
            DIAG_LOGGER.warning("工作表【%s】内容行数不足，有效数据为空或仅一行", sheet_name)
            return
        header = [str(cell or "").strip() for cell in rows[0]]
        col_idx = self._build_col_index(header, expected_cols)
        for row in rows[1:]:
            if not row:
                continue
            values = [str(cell or "").strip() if cell else "" for cell in row]
            yield col_idx, values

    def _read_l1_sheet(self, wb, config: LocationConfig):
        for col_idx, values in self._iter_sheet_rows(
            wb, self.L1_SHEET_NAME, ["服务器", "主机SN", "机房名称", "机柜编号", "L1名称", "L1_IP", "L1_SN"]
        ):
            host_name = values[col_idx["服务器"]] if "服务器" in col_idx else ""
            host_sn = values[col_idx["主机SN"]] if "主机SN" in col_idx else ""
            room_name = values[col_idx["机房名称"]] if "机房名称" in col_idx else ""
            cabinet_id = values[col_idx["机柜编号"]] if "机柜编号" in col_idx else ""
            l1_switch_name = values[col_idx["L1名称"]] if "L1名称" in col_idx else ""
            l1_switch_ip = values[col_idx["L1_IP"]] if "L1_IP" in col_idx else ""
            l1_switch_sn = values[col_idx["L1_SN"]] if "L1_SN" in col_idx else ""
            host_location_info = HostLocationInfo(
                host_name=host_name,
                host_sn=host_sn,
                room_name=room_name,
                cabinet_id=cabinet_id,
                l1_switch_name=l1_switch_name,
                l1_switch_ip=l1_switch_ip,
                l1_switch_sn=l1_switch_sn,
            )
            if host_name:
                config.host_name_location_map.update({host_name: host_location_info})
            if host_sn:
                config.host_sn_location_map.update({host_sn: host_location_info})
            switch_location_info = SwitchLocationInfo(
                sn=l1_switch_sn,
                room_name=room_name,
                cabinet_id=cabinet_id,
                switch_name=l1_switch_name,
                switch_ip=l1_switch_ip,
            )
            if l1_switch_sn:
                config.switch_sn_location_map.update({l1_switch_sn: switch_location_info})
            if l1_switch_ip:
                config.switch_ip_location_map.update({l1_switch_ip: switch_location_info})
            if l1_switch_name:
                config.switch_name_location_map.update({l1_switch_name: switch_location_info})

    def _read_l2_sheet(self, wb, config: LocationConfig):
        for col_idx, values in self._iter_sheet_rows(
            wb, self.L2_SHEET_NAME, ["设备名", "SN", "机房名称", "机柜编号", "管理IP配置"]
        ):
            switch_ip = values[col_idx["管理IP配置"]] if "管理IP配置" in col_idx else ""
            switch_name = values[col_idx["设备名"]] if "设备名" in col_idx else ""
            switch_sn = values[col_idx["SN"]] if "SN" in col_idx else ""
            switch_location_info = SwitchLocationInfo(
                sn=switch_sn,
                room_name=values[col_idx["机房名称"]] if "机房名称" in col_idx else "",
                cabinet_id=values[col_idx["机柜编号"]] if "机柜编号" in col_idx else "",
                switch_name=switch_name,
                switch_ip=switch_ip,
            )
            if switch_sn:
                config.switch_sn_location_map.update({switch_sn: switch_location_info})
            if switch_ip:
                config.switch_ip_location_map.update({switch_ip: switch_location_info})
            if switch_name:
                config.switch_name_location_map.update({switch_name: switch_location_info})
